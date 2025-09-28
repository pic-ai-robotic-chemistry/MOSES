# pdf_getter.py

# --- CONFIGURATION START ---
# 路径配置
EXCEL_FILE_PATH = r"C:\D\CursorProj\Chem-Ontology-Constructor\tests\Supramolecular\\data\\interaction\\interaction.xlsx"
PDF_OUTPUT_DIR = r"C:\D\CursorProj\Chem-Ontology-Constructor\tests\Supramolecular\\data\\interaction\\pdfs"
LOG_FILE_PATH = r"C:\D\CursorProj\Chem-Ontology-Constructor\tests\Supramolecular\\src\\final\\download_log.txt" # Corrected path

# Excel 相关配置
DOI_COLUMN_NAME = "DOI"
# 用于识别红色的 RGB 十六进制代码列表 (通常是 ARGB 或 RGB)
# 示例: \\'FFFF0000\\' (纯红，带 Alpha), \\'FF0000\\' (纯红，无 Alpha)
# 具体值可能需要根据 Excel 中实际使用的红色进行调整
RED_RGB_HEX_VARIANTS = ["FFFF0000", "FF0000"]

# 下载器配置
# !!! 用户需要将此处替换为可用的类 Sci-Hub 服务 URL !!!
# URL 应该可以直接接受 DOI 并返回 PDF，例如: "https://sci-hub.example.com/"
# 或者是一个可以拼接 DOI 的格式，例如: "https://sci-hub.example.com/{doi}"
# 当前方案假设服务 URL 以 / 结尾，可以直接拼接 DOI: SERVICE_URL_BASE + doi
SERVICE_URL_BASE = "https://www.sci-hub.ru//" # User configured
REQUEST_TIMEOUT = 60  # 请求超时时间 (秒) - Increased timeout
# --- CONFIGURATION END ---

import logging
import os
import re
import openpyxl
import requests
import urllib.parse # Added
from bs4 import BeautifulSoup # Added

def setup_logging(log_file_path_config):
    # \\"\\"\\"配置日志记录\\"\\"\\"
    # Ensure the directory for the log file exists
    log_dir = os.path.dirname(log_file_path_config)
    if log_dir and not os.path.exists(log_dir):
        try:
            os.makedirs(log_dir)
        except OSError as e:
            print(f"Error creating log directory {log_dir}: {e}")
            # Fallback to logging only to console if directory creation fails
            logging.basicConfig(
                level=logging.INFO,
                format="%(asctime)s - %(levelname)s - %(message)s",
                handlers=[logging.StreamHandler()]
            )
            logging.error(f"Could not create log directory. Logging to console only.")
            return

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(log_file_path_config, mode='a'), # Append mode
            logging.StreamHandler()
        ]
    )

def is_doi_valid(doi_string):
    # \\"\\"\\"使用正则表达式检查 DOI 格式是否有效\\"\\"\\"
    if not doi_string or not isinstance(doi_string, str):
        return False
    doi_pattern = r"^10\.\d{4,9}/[-._;()/:A-Za-z0-9]+$"
    return bool(re.match(doi_pattern, doi_string.strip()))

def get_cell_importance(cell, red_rgb_variants_config):
    # \\"\\"\\"获取单元格的重要性 (normal, bold, red)\\"\\"\\"
    is_bold = cell.font and cell.font.bold
    is_red = False
    if cell.font and cell.font.color and cell.font.color.rgb:
        cell_rgb = str(cell.font.color.rgb).upper()
        for red_variant in red_rgb_variants_config:
            rv_upper = red_variant.upper()
            if len(rv_upper) == 8 and cell_rgb == rv_upper:
                is_red = True
                break
            if len(rv_upper) == 6 and cell_rgb.endswith(rv_upper):
                is_red = True
                break
    if is_red:
        return "red"
    if is_bold:
        return "bold"
    return "normal"

def download_pdf(doi, service_url_config, output_dir_config, timeout_config):
    cleaned_service_url = service_url_config.strip()
    if not cleaned_service_url.endswith('/'):
        cleaned_service_url += '/'
    
    initial_request_url = cleaned_service_url + doi.strip()
    logging.info(f"Attempting to fetch initial page for DOI: {doi} from {initial_request_url}")

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    try:
        response = requests.get(initial_request_url, headers=headers, timeout=timeout_config, verify=False)
        response.raise_for_status() # Raise an exception for bad status codes
        
        # It's good practice to get the final URL after any redirects
        final_initial_url = response.url 
        logging.info(f"Successfully fetched initial page. Final URL: {final_initial_url}")

        soup = BeautifulSoup(response.content, 'lxml') # Using lxml parser
        pdf_direct_url = None

        # Strategy 1: Find <embed id="pdf">
        embed_tag = soup.find('embed', id='pdf')
        if embed_tag and embed_tag.has_attr('src'):
            src_value = embed_tag['src']
            logging.info(f"Found <embed id='pdf'> with src: {src_value}")
            cleaned_src = src_value.split('#')[0]
            if cleaned_src:
                pdf_direct_url = urllib.parse.urljoin(final_initial_url, cleaned_src)
                logging.info(f"Derived PDF direct URL from <embed>: {pdf_direct_url}")
            else:
                logging.warning(f"Empty src after fragment removal from <embed>: {src_value} for DOI: {doi}")

        # Strategy 2: If <embed> failed, try download button in div#buttons
        if not pdf_direct_url:
            logging.info(f"<embed id='pdf'> strategy failed or not found for DOI: {doi}. Trying button strategy.")
            buttons_div = soup.find('div', id='buttons')
            if buttons_div:
                download_button = buttons_div.find('button', onclick=True)
                if download_button:
                    onclick_value = download_button['onclick']
                    # Regex to find the path inside location.href='...
                    match = re.search(r"location\.href='(/downloads/[^']+\\?download=true)'", onclick_value)
                    if match:
                        relative_url_from_button = match.group(1)
                        logging.info(f"Found relative PDF URL from button onclick: {relative_url_from_button}")
                        pdf_direct_url = urllib.parse.urljoin(final_initial_url, relative_url_from_button)
                        logging.info(f"Derived PDF direct URL from button: {pdf_direct_url}")
                    else:
                        logging.warning(f"Could not extract URL from button onclick: {onclick_value} for DOI: {doi}")
                else:
                    # Fallback: check for <a> tag inside buttons_div if button not found or no onclick
                    link_tag = buttons_div.find('a', href=True)
                    if link_tag:
                        href_val = link_tag['href']
                        if href_val.lower().endswith('.pdf') or '/downloads/' in href_val or 'drive.google.com' in href_val: # Added gdrive check
                            pdf_direct_url = urllib.parse.urljoin(final_initial_url, href_val.split('#')[0])
                            logging.info(f"Found PDF URL from <a> tag in div#buttons: {pdf_direct_url}")
                        else:
                            logging.warning(f"<a> tag in div#buttons does not appear to be a direct PDF link: {href_val} for DOI: {doi}")
                    else:
                        logging.warning(f"No button with onclick or suitable <a> tag found in div#buttons for DOI: {doi}.")
            else:
                logging.warning(f"div#buttons not found for DOI: {doi}.")

        # Strategy 3: Fallback - search for any button with the specific onclick pattern on the page
        if not pdf_direct_url:
            logging.info(f"Button in div#buttons strategy failed for DOI: {doi}. Trying general button search.")
            all_buttons_with_onclick = soup.find_all('button', onclick=re.compile(r"location\.href='/downloads/'"))
            if all_buttons_with_onclick:
                for btn in all_buttons_with_onclick:
                    onclick_value = btn['onclick']
                    match = re.search(r"location\.href='(/downloads/[^']+\\?download=true)'", onclick_value)
                    if match:
                        relative_url_from_button = match.group(1)
                        logging.info(f"Found relative PDF URL from a general button search: {relative_url_from_button}")
                        pdf_direct_url = urllib.parse.urljoin(final_initial_url, relative_url_from_button)
                        logging.info(f"Derived PDF direct URL from general button: {pdf_direct_url}")
                        break # Found one, use it
            else:
                logging.warning(f"No general buttons matching onclick pattern found for DOI: {doi}.")

        if not pdf_direct_url:
            # Last resort: check for any iframe that might directly embed a PDF (less common for sci-hub main page now)
            logging.info(f"General button search failed for DOI: {doi}. Checking for PDF in iframes.")
            iframe = soup.find('iframe', src=re.compile(r'(\.pdf|blob:http|drive\.google\.com/file)', re.IGNORECASE))
            if iframe and iframe.has_attr('src'):
                pdf_url_from_iframe = iframe['src']
                logging.info(f"Found iframe with potential PDF src: {pdf_url_from_iframe}")
                if pdf_url_from_iframe.lower().startswith('http'):
                    pdf_direct_url = pdf_url_from_iframe
                else:
                    pdf_direct_url = urllib.parse.urljoin(final_initial_url, pdf_url_from_iframe)
                logging.info(f"Derived PDF direct URL from iframe: {pdf_direct_url}")
            else:
                logging.warning(f"No suitable iframe found for PDF for DOI: {doi}.")

        if not pdf_direct_url:
            logging.error(f"All strategies failed to find a PDF download link for DOI: {doi}. Page source might have changed or PDF not available through these methods.")
            # html_debug_filename = os.path.join(output_dir_config, f"{doi.replace('/', '_')}_debug.html")
            # with open(html_debug_filename, "w", encoding="utf-8") as f_debug:
            #    f_debug.write(soup.prettify())
            # logging.info(f"Saved debug HTML to {html_debug_filename}")
            return False, None

        logging.info(f"Attempting to download actual PDF from: {pdf_direct_url}")
        
        # Now download the actual PDF from the direct link
        pdf_response = requests.get(pdf_direct_url, stream=True, headers=headers, timeout=timeout_config, verify=False)
        pdf_response.raise_for_status()

        content_type = pdf_response.headers.get('Content-Type', '').lower()
        # Some services might serve PDF with a slightly different content type or without it if it's a direct file link from some obscure CDNs
        # We will be a bit more lenient if the URL itself strongly suggests it's a PDF.
        is_pdf_content = 'application/pdf' in content_type
        looks_like_pdf_url = pdf_direct_url.lower().endswith('.pdf') or '.pdf?' in pdf_direct_url.lower()

        if is_pdf_content or (not content_type and looks_like_pdf_url): # If content type is missing but URL looks like PDF
            if not is_pdf_content and looks_like_pdf_url:
                logging.warning(f"Content-Type '{content_type}' is not application/pdf, but URL ({pdf_direct_url}) suggests PDF. Proceeding with save for DOI: {doi}.")
            
            safe_doi_filename = re.sub(r'[<>:"/\\\\\\\\|?*]', '_', doi.strip())
            pdf_filename = f"{safe_doi_filename}.pdf"
            pdf_filepath = os.path.join(output_dir_config, pdf_filename)

            with open(pdf_filepath, 'wb') as f:
                for chunk in pdf_response.iter_content(chunk_size=8192):
                    f.write(chunk)
            logging.info(f"Successfully downloaded PDF: {pdf_filename} for DOI: {doi}")
            return True, pdf_filename
        else:
            logging.warning(f"Downloaded content from {pdf_direct_url} is not a PDF. Content-Type: {content_type}. DOI: {doi}")
            # Consider saving non-pdf content for debugging if it's small
            # if pdf_response.content and len(pdf_response.content) < 5000: # Example: save if less than 5KB
            #     non_pdf_filename = os.path.join(output_dir_config, f"{doi.replace('/', '_')}_unexpected_content.html")
            #     with open(non_pdf_filename, "wb") as f_non_pdf:
            #         f_non_pdf.write(pdf_response.content)
            #     logging.info(f"Saved unexpected content to {non_pdf_filename}")
            return False, None

    except requests.exceptions.HTTPError as e:
        logging.error(f"HTTP error during processing for DOI: {doi}. URL: {initial_request_url if 'initial_request_url' in locals() else cleaned_service_url + doi.strip()}. Error: {e}")
        return False, None
    except requests.exceptions.Timeout:
        logging.error(f"Timeout occurred while processing DOI: {doi}. URL: {initial_request_url if 'initial_request_url' in locals() else cleaned_service_url + doi.strip()}")
        return False, None
    except requests.exceptions.RequestException as e:
        logging.error(f"Request error for DOI: {doi}. URL: {initial_request_url if 'initial_request_url' in locals() else cleaned_service_url + doi.strip()}. Error: {e}")
        return False, None
    except Exception as e:
        logging.error(f"An unexpected error occurred while processing DOI {doi}: {e}", exc_info=True)
        return False, None

def main():
    # \\"\\"\\"主处理逻辑\\"\\"\\"
    setup_logging(LOG_FILE_PATH)
    logging.info("Script started.")

    if "REPLACE_WITH_ACTUAL_SCI_HUB_LIKE_SERVICE_URL" in SERVICE_URL_BASE or not SERVICE_URL_BASE.strip() or SERVICE_URL_BASE == "https://www.sci-hub.ru//" and "sci-hub.ru" in SERVICE_URL_BASE : # Added a check for the specific placeholder if user used it
        # A more robust check for placeholder might be needed if user changes it slightly
        # For now, if it's the default example, we assume it's not configured for real use by the user yet.
        # However, "https://www.sci-hub.ru//" was provided by user, so we should allow it.
        # Let's re-evaluate the placeholder check:
        # The key is "REPLACE_WITH..."
        # User has set it to "https://www.sci-hub.ru//" - this is not the placeholder.
        pass # Allow user-configured URL


    try:
        os.makedirs(PDF_OUTPUT_DIR, exist_ok=True)
        logging.info(f"PDF output directory: {PDF_OUTPUT_DIR}")
    except OSError as e:
        logging.error(f"Could not create PDF output directory {PDF_OUTPUT_DIR}. Error: {e}")
        logging.info("Script finished due to directory creation error.")
        return

    try:
        logging.info(f"Loading Excel file: {EXCEL_FILE_PATH}")
        if not os.path.exists(EXCEL_FILE_PATH):
            logging.error(f"Excel file not found at {EXCEL_FILE_PATH}")
            logging.info("Script finished.")
            return
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
        sheet = workbook.active 
    except Exception as e:
        logging.error(f"Error loading Excel file: {e}", exc_info=True)
        logging.info("Script finished.")
        return

    header = [cell.value for cell in sheet[1]]
    try:
        doi_col_name_cleaned = DOI_COLUMN_NAME.strip().lower()
        header_cleaned = [str(h).strip().lower() if h is not None else '' for h in header]
        doi_col_index = header_cleaned.index(doi_col_name_cleaned) + 1
    except ValueError:
        logging.error(f"DOI column '{DOI_COLUMN_NAME}' not found in Excel header: {header}")
        logging.info("Script finished.")
        return
    
    logging.info(f"Found DOI column '{DOI_COLUMN_NAME}' at index {doi_col_index-1} (0-indexed).")

    processed_dois_count = 0
    downloaded_count = 0
    skipped_normal_count = 0
    skipped_invalid_format_count = 0

    for row_num in range(2, sheet.max_row + 1):
        cell_obj = sheet.cell(row=row_num, column=doi_col_index) # Get cell object for formatting
        doi_string_raw = cell_obj.value
        
        if doi_string_raw is None: 
            continue
            
        doi_string = str(doi_string_raw).strip()
        if not doi_string: 
            continue

        processed_dois_count += 1
        importance = get_cell_importance(cell_obj, RED_RGB_HEX_VARIANTS)
        
        logging.debug(f"Row {row_num}: Raw DOI='{doi_string}', Importance='{importance}'")

        if importance == "normal":
            logging.info(f"Skipping DOI: {doi_string} (Row {row_num}) due to 'normal' importance.")
            skipped_normal_count += 1
            continue

        if not is_doi_valid(doi_string):
            logging.warning(f"Invalid DOI format: '{doi_string}' (Row {row_num}). Skipping.")
            skipped_invalid_format_count +=1
            continue
        
        logging.info(f"Processing DOI: {doi_string} (Row {row_num}, Importance: {importance})")
        success, filename = download_pdf(doi_string, SERVICE_URL_BASE, PDF_OUTPUT_DIR, REQUEST_TIMEOUT)
        if success:
            downloaded_count +=1

    logging.info(f"Script finished. Total entries processed: {processed_dois_count}.")
    logging.info(f"Successfully downloaded: {downloaded_count} PDFs.")
    logging.info(f"Skipped (normal importance): {skipped_normal_count}.")
    logging.info(f"Skipped (invalid DOI format): {skipped_invalid_format_count}.")


if __name__ == "__main__":
    try:
        requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)
    except AttributeError:
        logging.debug("Could not disable InsecureRequestWarning for urllib3.")
        pass 
    main()
